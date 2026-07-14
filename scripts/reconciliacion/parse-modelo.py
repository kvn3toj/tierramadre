# scripts/reconciliacion/parse-modelo.py — SOLO LECTURA
#
# Reconciliación Fase 1, Task 1: parsea la hoja "Inventario" del Modelo de
# fijación de precios (xlsx descargado vía Google Drive MCP, ver fetch.ts /
# README del task) a JSON local bajo out/. No escribe a Drive/Sheets.
#
# Uso: python3 scripts/reconciliacion/parse-modelo.py out/modelo.xlsx
import openpyxl, json, sys, os

wb = openpyxl.load_workbook(sys.argv[1], data_only=True)  # sheet.xlsx
ws = wb['Inventario']


def fixnum(v):
    v = str(v).strip() if v is not None else ''
    return v[:-2] if v.endswith('.0') else v


rows = []
for r in range(2, ws.max_row + 1):
    cod, item = ws.cell(r, 1).value, ws.cell(r, 3).value
    if cod is None and item is None:
        continue
    rows.append({
        'codigo': str(cod).strip() if cod else '',
        'item': fixnum(item),
        'corte': str(ws.cell(r, 4).value or '').strip(),
        'unid': ws.cell(r, 5).value,
        'nombreLote': str(ws.cell(r, 6).value or '').strip(),
        'calidad': str(ws.cell(r, 9).value or '').strip(),
        'costo': ws.cell(r, 11).value,
    })

out_dir = 'scripts/reconciliacion/out'
os.makedirs(out_dir, exist_ok=True)
json.dump(rows, open(os.path.join(out_dir, 'modelo.json'), 'w'), ensure_ascii=False, indent=0)
print('modelo rows:', len(rows), '| con item#:', sum(1 for x in rows if x['item']))
